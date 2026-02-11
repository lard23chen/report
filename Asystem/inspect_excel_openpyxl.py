
from openpyxl import load_workbook

file_path = "d:\\2025\\AI\\MongoDB\\Asystem\\asys_aws_移行log改寫清單.xlsx"

try:
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = row
    
    print("Columns:", headers)
except Exception as e:
    print(f"Error reading file: {e}")
