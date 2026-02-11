
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

file_path = "d:\\2025\\AI\\MongoDB\\Asystem\\asys_aws_移行log改寫清單.xlsx"
summary_sheet_name = "負責人總表分析"

try:
    print(f"Loading workbook: {file_path}...")
    wb = openpyxl.load_workbook(file_path)
    
    # Assume the first sheet is the data source
    data_sheet = wb.worksheets[0]
    print(f"Reading data from sheet: {data_sheet.title}")
    
    # Headers are in row 1
    headers = [cell.value for cell in data_sheet[1]]
    try:
        person_idx = headers.index("負責人")
        date_idx = headers.index("完成日")
    except ValueError:
        print("Error: Could not find '負責人' or '完成日' columns in the active sheet.")
        exit(1)

    stats = {} # person -> {total, completed, pending}

    print("Analyzing data...")
    for row in data_sheet.iter_rows(min_row=2, values_only=True):
        person = row[person_idx]
        completed_date = row[date_idx]
        
        person_key = str(person).strip() if person else "未指定"
        
        if person_key not in stats:
            stats[person_key] = {"total": 0, "completed": 0, "pending": 0}
        
        stats[person_key]["total"] += 1
        if completed_date:
            stats[person_key]["completed"] += 1
        else:
            stats[person_key]["pending"] += 1

    # Prepare summary data
    summary_data = []
    # Header
    summary_data.append(["負責人", "總任務數", "已完成", "未完成", "完成率"])
    
    # Sort by total tasks descending
    sorted_people = sorted(stats.items(), key=lambda item: item[1]['total'], reverse=True)
    
    total_row = {"total": 0, "completed": 0, "pending": 0}

    for person, counts in sorted_people:
        rate = (counts['completed'] / counts['total']) if counts['total'] > 0 else 0
        summary_data.append([
            person,
            counts['total'],
            counts['completed'],
            counts['pending'],
            f"{rate:.1%}"
        ])
        total_row['total'] += counts['total']
        total_row['completed'] += counts['completed']
        total_row['pending'] += counts['pending']

    # Add Total Row
    total_rate = (total_row['completed'] / total_row['total']) if total_row['total'] > 0 else 0
    summary_data.append([
        "總計",
        total_row['total'],
        total_row['completed'],
        total_row['pending'],
        f"{total_rate:.1%}"
    ])

    # Manage Summary Sheet
    if summary_sheet_name in wb.sheetnames:
        print(f"Removing existing '{summary_sheet_name}' sheet...")
        std = wb[summary_sheet_name]
        wb.remove(std)
    
    print(f"Creating new '{summary_sheet_name}' sheet...")
    ws_summary = wb.create_sheet(title=summary_sheet_name)
    
    # Write data
    for row in summary_data:
        ws_summary.append(row)

    # Formatting
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Format Header
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # Format Data
    for row in ws_summary.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border

    # Auto-adjust column width
    for col in ws_summary.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)
        ws_summary.column_dimensions[column].width = adjusted_width

    wb.save(file_path)
    print(f"Successfully updated {file_path}")

except PermissionError:
    print(f"Permission denied: Please close '{file_path}' and try again.")
except Exception as e:
    print(f"An error occurred: {e}")
