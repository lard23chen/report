
from openpyxl import load_workbook

file_path = "d:\\2025\\AI\\MongoDB\\Asystem\\asys_aws_移行log改寫清單.xlsx"

try:
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    
    print("First 5 rows of completion dates:")
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True)):
        print(f"Row {i+2}: {row[10]}") # Index 10 is '完成日'

except Exception as e:
    print(f"Error: {e}")
