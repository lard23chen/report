
from openpyxl import load_workbook
from collections import defaultdict

file_path = "d:\\2025\\AI\\MongoDB\\Asystem\\asys_aws_移行log改寫清單.xlsx"

try:
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    
    # Get header row to find column indices
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = row
        break
    
    try:
        person_col_idx = headers.index("負責人")
        date_col_idx = headers.index("完成日")
    except ValueError:
        print("Could not find '負責人' or '完成日' column")
        exit()

    stats = defaultdict(lambda: {"total": 0, "completed": 0, "pending": 0})
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        person = row[person_col_idx]
        completion_date = row[date_col_idx]
        
        if person is None:
            person = "未指定 (Unassigned)"
        else:
            person = str(person).strip()
            
        stats[person]["total"] += 1
        
        if completion_date:
            stats[person]["completed"] += 1
        else:
            stats[person]["pending"] += 1

    print("\n負責人任務統計分析 (Person in Charge Task Analysis):")
    print("-" * 60)
    print(f"{'負責人':<20} | {'總任務數':<10} | {'已完成':<10} | {'未完成':<10} | {'完成率':<10}")
    print("-" * 60)
    
    for person, data in stats.items():
        total = data["total"]
        completed = data["completed"]
        pending = data["pending"]
        rate = (completed / total * 100) if total > 0 else 0
        
        print(f"{person:<20} | {total:<10} | {completed:<10} | {pending:<10} | {rate:.1f}%")

except Exception as e:
    print(f"Error analyzing file: {e}")
